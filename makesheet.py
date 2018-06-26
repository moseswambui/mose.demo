import openpyxl

worksheet= openpyxl.load_workbook('Book1.xlsx')
#mysheet= worksheet.get_sheet_names()
print(worksheet.sheetnames)
print(worksheet.get_active_sheet)
use_sheet= worksheet.sheetnames
print(use_sheet)
loading=worksheet.get_active_sheet
print(loading)
wb=worksheet['Sheet1']
print(wb)
print(wb['A1'].value)
print(wb.cell(row=3, column=4).value)
print(tuple(wb['A1' :'C3']))