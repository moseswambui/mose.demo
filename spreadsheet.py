import openpyxl

load=openpyxl.load_workbook('Book1.xlsx')
sheet=load['Sheet1']
print(sheet['A1'].value)
active= load.active
print(active)
b= sheet['B2']
print(b.value)
print(sheet.cell(row=1, column=2).value)
for i in range(1,8):
    print(i, sheet.cell(row=i, column=3).value)